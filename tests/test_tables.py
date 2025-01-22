"""Tests the creation of the tables."""

import logging
import os

import openpyxl
import pandas as pd


if __name__ == "__main__":
    dirname = os.path.dirname
    os.chdir(path=os.path.join(dirname(dirname(__file__))))

from analysis import tables


logger = logging.getLogger(__name__)


def test_create_pivot_table(
    input_create_pivot_table: pd.DataFrame, output_create_pivot_table: pd.DataFrame
) -> None:
    """Tests whether the pivot table is created as expected based on input DataFrame.

    Args:
        input_create_pivot_table (pd.DataFrame): Fixture with the input for the pivot table.
        output_create_pivot_table (pd.DataFrame): Fixture with the expected output as pivot table.
    """
    result = tables.create_pivot_table(
        input_create_pivot_table,
        "Gebied",
        "Margalef_Monster",
        ["Extra_kolom_1", "Extra_kolom_2"],
    )
    result = result.rename_axis(None, axis=1)

    pd.testing.assert_frame_equal(
        result.reset_index(drop=True),
        output_create_pivot_table.reset_index(drop=True),
    )


def test_get_min_year() -> None:
    """Tests whether the right minimum year is returned from the configuration."""
    waterbodies = ["Waddenzee", "Noordzee"]
    output = 1987

    result = tables.get_min_year(waterbodies)
    assert result == output


def test_export_samples_a_year_basic(
    input_sample_a_year: pd.DataFrame, output_sample_a_year: pd.DataFrame
) -> None:
    """Tests whether the samples a year given the result as expected.

    Args:
        input_sample_a_year (pd.DataFrame): Fixture with input benthos data.
        output_sample_a_year (pd.DataFrame): Fixture with the expected samples a year.
    """
    result = tables.export_samples_a_year(input_sample_a_year)
    result = result.rename_axis(None, axis=1)
    result.columns = result.columns.astype(str)

    pd.testing.assert_frame_equal(
        result.reset_index(drop=True),
        output_sample_a_year.reset_index(drop=True),
    )

    assert os.path.isfile("./output/Monsters_per_jaar.xlsx")
    excel = pd.read_excel("./output/Monsters_per_jaar.xlsx")
    excel.columns = excel.columns.astype(str)
    excel = excel.astype(object)
    output_sample_a_year = output_sample_a_year.astype(object)

    pd.testing.assert_frame_equal(
        excel.reset_index(drop=True),
        output_sample_a_year.reset_index(drop=True),
    )


def test_export_samples_a_year_main(
    input_analysis_main: pd.DataFrame, output_analysis_sample_a_year: pd.DataFrame
) -> None:
    """Tests whether the main import produces the right samples a year data.

    Args:
        input_analysis_main (pd.DataFrame): Fixture with the input for the main integration function.
        output_analysis_sample_a_year (pd.DataFrame): Fixture with the output of the samples a year.
    """
    result = tables.export_samples_a_year(input_analysis_main)
    result = result.rename_axis(None, axis=1)
    result.columns = result.columns.astype(str)

    result = result.astype(object)
    expected = output_analysis_sample_a_year.astype(object)

    pd.testing.assert_frame_equal(
        result.reset_index(drop=True),
        expected.reset_index(drop=True),
    )

    assert os.path.isfile("./output/Monsters_per_jaar.xlsx")
    excel = pd.read_excel("./output/Monsters_per_jaar.xlsx")
    excel.columns = excel.columns.astype(str)
    excel = excel.astype(object)

    pd.testing.assert_frame_equal(
        excel.reset_index(drop=True),
        expected.reset_index(drop=True),
    )


def compare_excel_files(file1: str, file2: str) -> None:
    """Compares to excel files based on sheets and rows with values.

    Args:
        file1 (str): the file path of one file.
        file2 (str): the file path of the other file.
    """
    workbook1 = openpyxl.load_workbook(file1)
    workbook2 = openpyxl.load_workbook(file2)

    # Get the list of sheet names from both workbooks
    sheet_names1 = workbook1.sheetnames

    # Iterate through sheets and compare the cell values
    for sheet_name in sheet_names1:
        sheet1 = workbook1[sheet_name]
        sheet2 = workbook2[sheet_name]

        # Check if the sheets have the same number of rows and columns
        assert (
            sheet1.max_row == sheet2.max_row
        ), f"Number of rows is different in sheet: {sheet_name}"
        assert (
            sheet1.max_column == sheet2.max_column
        ), f"Number of columns is different in sheet: {sheet_name}"

        for _, (row1, row2) in enumerate(
            zip(sheet1.iter_rows(), sheet2.iter_rows()), start=1
        ):
            for _, (cell1, cell2) in enumerate(zip(row1, row2), start=1):
                if cell1.value != cell2.value:
                    assert (
                        False  # This line stops the comparison on the first difference
                    )


def test_make_species_list_wadkust(
    input_species_list_wadkust: pd.DataFrame,
    output_species_list_wadkust: pd.DataFrame,
    habitat_species: pd.DataFrame,
    mocker: pd.DataFrame,
) -> None:
    """Tests whether the species list of the Waddenkust is as expected incl. habitat species.

    Args:
        input_species_list_wadkust (pd.DataFrame): Fixture with the input species.
        output_species_list_wadkust (pd.DataFrame): Fixture with the expected output.
        habitat_species (pd.DataFrame): Fixture with the the n2000 habitat species.
        mocker (pd.DataFrame): The output of the check_habitat_n2000_species_conform_twn is habitat_species.
    """
    mocker.patch(
        "checks.check_tables.check_habitat_n2000_species_conform_twn",
        return_value=habitat_species,
    )
    result = tables.make_species_list(input_species_list_wadkust)

    expected = output_species_list_wadkust

    pd.testing.assert_frame_equal(
        result.reset_index(drop=True),
        expected.reset_index(drop=True),
    )


def test_make_species_list_Noordzee(
    input_species_list_noordzee: pd.DataFrame,
    output_species_list_noordzee: pd.DataFrame,
    habitat_species: pd.DataFrame,
    mocker: pd.DataFrame,
) -> None:
    """Tests whether the species list of the Noordzee is as expected incl. habitat species.

    Args:
        input_species_list_wadkust (pd.DataFrame): Fixture with the input species.
        output_species_list_wadkust (pd.DataFrame): Fixture with the expected output.
        habitat_species (pd.DataFrame): Fixture with the the n2000 habitat species.
        mocker (pd.DataFrame): The output of the check_habitat_n2000_species_conform_twn is habitat_species.
    """
    mocker.patch(
        "checks.check_tables.check_habitat_n2000_species_conform_twn",
        return_value=habitat_species,
    )
    result_noordzee = tables.make_species_list(input_species_list_noordzee)

    pd.testing.assert_frame_equal(
        result_noordzee.reset_index(drop=True),
        output_species_list_noordzee.reset_index(drop=True),
    )
