"""The Benthic Indicator Species Index (BISI) will be calculated."""

"""
# File: BISI.py
# Author(s): M. Japink (m.japink@waardenburg.eco) & J. Haringa (j.haringa@waardenburg.eco)
# Creation date: 21-01-2023
# Last modification: 20-02-2024
# Python v3.12.1
"""

import logging
import os
import shutil

from openpyxl import load_workbook
import pandas as pd


if __name__ == "__main__":
    dirname = os.path.dirname
    os.chdir(path=os.path.join(dirname(dirname(__file__))))

from preparation import log_decorator
from preparation import process_twn
from preparation import read_system_config


logger = logging.getLogger(__name__)


@log_decorator.log_factory(__name__)
def read_bisi_criteria(
    output_path: str, bisi_sheet_sheetname: str, cell_index: int
) -> pd.DataFrame:
    """Read the BISI criteria from the BISI sheet for a given area.
    So minor issues are fixed in the BISI criteria.

    Args:
        output_path (str): The path to the BISI sheet.
        bisi_sheet_sheetname (str): The sheetname of the BISI sheet.
        cell_index (int): The index of the cell where the BISI criteria start.

    Returns:
        pd.DataFrame: The dataframe with the corrected BISI criteria.
    """

    # retrieve the indicator species (column A), sampling techniques (H),
    # and exp. nr of samples (K) from the BISI sheet
    df_bisi_criteria = pd.read_excel(
        output_path,
        sheet_name=bisi_sheet_sheetname,
        skiprows=cell_index,
        usecols="A, H, K",
    )

    # find the first empty row to retrieve the indicator species
    first_empty_row = df_bisi_criteria[
        df_bisi_criteria.isnull().all(axis=1)
    ].index.tolist()[0]
    df_bisi_criteria = df_bisi_criteria.loc[0 : first_empty_row - 1]
    df_bisi_criteria.columns = [
        "Analyse_taxonnaam",
        "Bemonsteringsapp",
        "Expected_n",
    ]

    # add a position column to the dataframe to keep the order of the taxa
    df_bisi_criteria["Position"] = df_bisi_criteria.index + 1

    # fix bemonsteringsapparaat
    # fix missing spaces before the m2
    df_bisi_criteria["Bemonsteringsapp"] = df_bisi_criteria["Bemonsteringsapp"].apply(
        lambda x: add_missing_space_before(x, "m2")
    )

    # fix N_Expected
    # fill N_expected Nan with 0
    df_bisi_criteria["Expected_n"].fillna(0, inplace=True)
    # convert N_expected to int
    df_bisi_criteria["Expected_n"] = df_bisi_criteria["Expected_n"].astype(int)
    return df_bisi_criteria


@log_decorator.log_factory(__name__)
def add_missing_space_before(input_string: str, post_characters: str) -> str:
    """Add a space before a specified character if there is not already a space before it.

    Args:
        input_string (str): The string to check.
        post_characters (str): The character to check the position direct before.

    Returns:
        str: The string with a space before the specified character.
    """
    # checking in all lower case
    lower_string = input_string.lower()
    lower_characters = post_characters.lower()

    space_position = lower_string.find(lower_characters)
    if space_position <= 0:
        return input_string

    # Fixing on original string
    if input_string[space_position - 1] == " ":
        return input_string
    return input_string[:space_position] + " " + input_string[space_position:]


@log_decorator.log_factory(__name__)
def check_bisi_taxa(df_bisi_criteria: pd.DataFrame, bisi_area: str) -> bool:
    """Check the taxa in the BISI sheet.
    There might be three problems with the BISI taxa:
    - they are not in the twn or
    - they are invalid without an synonym or
    - they are invalid with a synonym
    In all three cases we return a warning. Those taxa are not included in the BISI sheet.
    We are not going to fix the taxa in the BISI sheet, because it should be common species.
    If TWN changes on any point, we should change the BISI sheet.

    Args:
        df_bisi_criteria (pd.DataFrame): Dataframe with the BISI species.

    Returns:
        bool: True if all taxa in the BISI sheet are valid, False otherwise.
    """

    twn_corrected = read_system_config.read_csv_file(
        read_system_config.read_yaml_configuration(
            "twn_corrected", "global_variables.yaml"
        )
    )

    df_bisi_species = df_bisi_criteria.copy()

    # fix the abbreviated genus names
    df_bisi_species["Analyse_taxonnaam"] = df_bisi_species["Analyse_taxonnaam"].apply(
        lambda x: fix_abbreviated_genus_names(x)
    )

    # split combined species to separate rows
    df_bisi_species["Analyse_taxonnaam"] = df_bisi_species[
        "Analyse_taxonnaam"
    ].str.split(" \\+ ")
    df_bisi_species = df_bisi_species.explode("Analyse_taxonnaam")

    df_bisi_species["Analyse_taxonnaam"] = df_bisi_species["Analyse_taxonnaam"].apply(
        lambda x: remove_taxa_postfixes(x)
    )

    # get all taxa statusses from the twn
    twn_validity = process_twn.get_twn_validity(twn_corrected, df_bisi_species)
    # check uknown or invalid taxa
    check_invalid = twn_validity[twn_validity["Status"].isin(["unknown", "invalid"])]
    if not check_invalid.empty:
        logger.warning(
            f"Voor {bisi_area} zijn de volgende taxa in de BISI tabel onbekend of ongeldig in de TWN."
            f'Ze worden dus niet meegenomen: \n{check_invalid["Analyse_taxonnaam"].unique()}'
        )

    check_synonym = twn_validity[twn_validity["Status"].isin(["synonym"])]
    if not check_synonym.empty:
        logger.warning(
            f"Voor {bisi_area} zijn de volgende taxa in de BISI ongeldig en hebben een voorkeursnaam in de TWN."
            "Ze worden niet meegenomen. Controleer en pas waar nodig de BISI aan:"
            f'\n{check_synonym["Analyse_taxonnaam", "Synonymname"].drop_duplicates()}'
        )

    if not check_invalid.empty or not check_synonym.empty:
        return False
    logger.info("Alle taxa in de BISI tabel zijn bekend en geldig in de TWN.")
    return True


def check_required_area(df_bisi: pd.DataFrame, bisi_col: str) -> bool:
    """Check whether the required area for the BISI index is the same in the BISI sheet and the data.
    Only warn the user if the area is not the same.

    Args:
        df_bisi (pd.DataFrame): dataframe with collected data for the BISI index.
        bisi_col (str): the column name of the BISI index.

    Returns:
        bool: True if the area is the same in the BISI sheet and the data, False otherwise.
    """

    bisi_gebied = df_bisi[bisi_col].unique()
    df_bisi = df_bisi.copy()
    df_bisi["Bemonsteringsapp"] = df_bisi["Bemonsteringsapp"].str.replace(
        ",", ".", regex=False
    )
    df_bisi = df_bisi[df_bisi["Support_Eenheid"] == "m2"]
    df_bisi["Support"] = df_bisi["Support"].astype(str)
    df_bisi["Support"] = df_bisi["Support"].str.replace(".0$", "", regex=True)
    df_bisi["Compare_opp"] = df_bisi.apply(
        lambda x: str(x["Support"]) in str(x["Bemonsteringsapp"]), axis=1
    )

    if df_bisi["Compare_opp"].all():
        logger.debug(f"Correct area for the BISI index for {bisi_gebied}.")
        return True
    logger.warning(
        f"De bemonsterde oppervlaktes in de BISI rekensheet en de data zijn niet gelijk in "
        f"{bisi_col}:{bisi_gebied}: \n "
        f'{df_bisi[df_bisi["Compare_opp"] == False][["Support", "Bemonsteringsapp"]].drop_duplicates()}'
    )
    return False


@log_decorator.log_factory(__name__)
def check_sample_species(
    df: pd.DataFrame, df_bisi_criteria: pd.DataFrame, bisi_col: str
) -> pd.DataFrame:
    """Calculate the number of samples for each area per year and checks which
    species are in the BISI Excel and compares this to do BISI data. If the samples are to low or
    species are not present, the user is given a warning in the logfile.

    Args:
        df (pd.DataFrame): the DataFrame with selected Aqaudesk data
        df_bisi_criteria (pd.DataFrame): dataframe with the bisi data from the BISI excel sheet.
        bisi_col (str): the column name of the BISI index.

    Returns:
        pd.DataFrame: dataframe with the aggregated number of samples for each area and year.
    """

    bisi_gebied = df[bisi_col].unique()

    df_monsters = (
        df.groupby(
            ["Monsterjaar", bisi_col, "Bemonsteringsapp"], dropna=False, as_index=False
        )
        .agg(Nmonsters=("Collectie_Referentie", "nunique"))
        .reset_index()
    )

    # check required number of samples
    try:
        for bem_app in df_monsters["Bemonsteringsapp"].unique():
            samples = df_monsters[df_monsters["Bemonsteringsapp"].isin([bem_app])][
                "Nmonsters"
            ].unique()
            expected_samples = df_bisi_criteria[
                df_bisi_criteria["Bemonsteringsapp"].isin([bem_app])
            ]["Expected_n"].unique()
            if samples < expected_samples:
                logger.warning(
                    f"Het aantal monsters ({samples}) is minder dan het verwachte aantal"
                    f"({expected_samples}) voor {bem_app} in {bisi_gebied}."
                    "Resultaten kunnen daardoor minder representatief zijn!"
                )
    except Exception:
        logger.warning(
            f"Er kan geen check gedaan worden voor het aantal monsters voor {bisi_gebied}."
        )

    # check required species
    check = df_bisi_criteria.loc[
        ~df_bisi_criteria["Analyse_taxonnaam"].isin(df["Analyse_taxonnaam"]),
        "Analyse_taxonnaam",
    ]
    if check.any():
        logger.warning(
            f"Niet alle BISI-taxa zijn aanwezig voor {bisi_gebied}:\n {check}"
        )

    return df_monsters


def map_sampling_device_to_bisi(
    df: pd.DataFrame, bisi_sheet_sampling_devices_dict: dict
) -> pd.DataFrame:
    """Map the sampling devices according to the BISI sheet (by config)

    Args:
        df (pd.DataFrame): The dataframe with the Aquadesk data.
        bisi_sheet_sampling_devices_dict (dict): config dictionary with the sampling devices.

    Returns:
        pd.DataFrame: The dataframe with the Aquadesk data with the sampling devices according to the BISI sheet.
    """

    for key, value in bisi_sheet_sampling_devices_dict.items():
        df.loc[df["Bemonsteringsapp"].isin(value.split("/")), "Bemonsteringsapp"] = key

    return df


def remove_taxa_postfixes(taxa_string: str) -> str:
    """Fix all taxa postfixes in the BISI table.

    Args:
        taxa_string (str): The taxa string.

    Returns:
        str: The taxa string with removed postfixes.
    """

    taxa_string = taxa_string.replace("*", "")
    taxa_string = taxa_string.replace("**", "")
    taxa_string = taxa_string.replace(" spp.", "")
    return taxa_string


def fix_abbreviated_genus_names(taxa_string: str) -> str:
    """Fix abbreviated genus names in the BISI table with the first word of the taxa-string.

    Args:
        taxastring (str): The taxa string with correct and abbreviated genus names.

    Returns:
        str: The taxa string with correct genus names.
    """
    genus_name = taxa_string.split()[0]
    full_names = taxa_string.replace(genus_name[:1] + ".", genus_name)
    return full_names


def map_taxa_to_bisi(df: pd.DataFrame, df_bisi_criteria: pd.DataFrame) -> pd.DataFrame:
    """Map the taxonames in the data to the BISI taxonames.

    Args:
        df (pd.DataFrame): The dataframe with the Aquadesk data.
        df_bisi_criteria (pd.DataFrame): The dataframe with the BISI taxa.


    Returns:
        pd.DataFrame: The dataframe with the Aquadesk data mapped to the BISI taxa.
    """

    df_bisi_criteria = df_bisi_criteria.copy()

    # copy the taxon column and split the combined species to separate rows
    df_bisi_criteria["BISI_taxonnaam"] = df_bisi_criteria["Analyse_taxonnaam"]

    # fix the abbreviated genus names
    df_bisi_criteria["Analyse_taxonnaam"] = df_bisi_criteria["Analyse_taxonnaam"].apply(
        lambda x: fix_abbreviated_genus_names(x)
    )

    df_bisi_criteria["Analyse_taxonnaam"] = df_bisi_criteria[
        "Analyse_taxonnaam"
    ].str.split(" \\+ ")
    df_bisi_criteria = df_bisi_criteria.explode("Analyse_taxonnaam")

    # fix all other bisi taxa postfixes
    df_bisi_criteria["Analyse_taxonnaam"] = df_bisi_criteria["Analyse_taxonnaam"].apply(
        lambda x: remove_taxa_postfixes(x)
    )
    spp_taxa_dict = df_bisi_criteria.set_index("Analyse_taxonnaam")[
        "BISI_taxonnaam"
    ].to_dict()

    # store original name for logging
    df["Analyse_taxonnaam_orig"] = df["Analyse_taxonnaam"]

    # replace the taxa names by the BISI taxa names
    for key, value in spp_taxa_dict.items():
        df.loc[
            (df["Analyse_taxonnaam"].isin([key]))
            | (df["Hierarchie"].str.contains(key)),
            "Analyse_taxonnaam",
        ] = value

    # log the mapped taxa names for this BISI area
    df_taxa_mapping = df[
        ["Analyse_taxonnaam_orig", "Analyse_taxonnaam"]
    ].drop_duplicates()
    df_taxa_mapping = df_taxa_mapping[
        df_taxa_mapping["Analyse_taxonnaam_orig"]
        != df_taxa_mapping["Analyse_taxonnaam"]
    ]
    df_taxa_mapping.rename(
        columns={
            "Analyse_taxonnaam_orig": "Analyse_taxonnaam",
            "Analyse_taxonnaam": "BISI_taxonnaam",
        },
        inplace=True,
    )

    # clean up the data
    df.drop(columns=["Analyse_taxonnaam_orig"], inplace=True)

    return df


def bisi_calculations(df: pd.DataFrame, df_bisi_criteria: pd.DataFrame) -> pd.DataFrame:
    """Calculate the BISI values for every taxon. A few steps are performed:
    - for each sample the missing taxa are padded with 0
    - the samples are summed by taxon, ann sampling device
    - the average density is calculated for each taxon and sampling device
    - the standard deviation is calculated for each taxon and sampling device

    Args:
        df (pd.DataFrame): dataframe with the Aquadesk data.
        df_bisi_criteria: dataframe with bis criteria

    Returns:
        pd.DataFrame: The dataframe with the BISI values.
    """

    # get datafram with 1 column of unique sample codes (Collectie_Referentie)
    df_samples = df["Collectie_Referentie"].drop_duplicates().to_frame()

    # cross join the samples with the bisi_criteria
    df_samples["key"] = 1
    df_bisi_criteria["key"] = 1
    df_filler = pd.merge(df_samples, df_bisi_criteria, on="key", how="left").drop(
        "key", axis=1
    )

    # join only necessary sample information to the df_filler
    df_calc = pd.merge(
        df_filler,
        df,
        on=["Collectie_Referentie", "Analyse_taxonnaam", "Bemonsteringsapp"],
        how="left",
    )
    df_calc = df_calc[
        [
            "Collectie_Referentie",
            "Analyse_taxonnaam",
            "Bemonsteringsapp",
            "Dichtheid_Aantal",
            "Position",
        ]
    ]

    df_calc["Dichtheid_Aantal"].fillna(0, inplace=True)

    # sum within sample
    df_calc_sum = df_calc.groupby(
        ["Collectie_Referentie", "Analyse_taxonnaam", "Bemonsteringsapp", "Position"],
        dropna=False,
        as_index=False,
    ).aggregate(
        {"Dichtheid_Aantal": "sum"},
    )

    # calculate the average and standard deviation for each taxon
    df_calc_sum["Stdev"] = df_calc_sum["Dichtheid_Aantal"]
    df_calc_agg = df_calc_sum.groupby(
        ["Analyse_taxonnaam", "Bemonsteringsapp", "Position"],
        dropna=False,
        as_index=False,
    ).aggregate({"Dichtheid_Aantal": "mean", "Stdev": "std"})

    # round the average density and standard deviation
    df_calc_agg["Dichtheid_Aantal"] = df_calc_agg["Dichtheid_Aantal"].round(6)
    df_calc_agg["Stdev"] = df_calc_agg["Stdev"].round(6)

    # fill NAs with 0
    df_calc_agg["Dichtheid_Aantal"].fillna(0, inplace=True)
    df_calc_agg["Stdev"].fillna(0, inplace=True)

    return df_calc_agg


@log_decorator.log_factory(__name__)
def main_bisi(df: pd.DataFrame, year: int = None) -> pd.DataFrame:
    """Retrieve the BISI-gebied, connects this to a sheet in the BISI table,
    checks whether the species in the BISI table are also found for a minimal number of samples
    in the data. If so, it writes the number of n for this species in the BISI table.

    Args:
        df (pd.DataFrame): The dataframe with the data.
        year (int, optional): The year for which the BISI should be calculated. Defaults to 2018.

    Returns:
        pd.DataFrame: The dataframe with the BISI values of the last processed area.

    Raises:
        Exception: If there are not enough samples for one or more species, the BISI table should not
        be filled in because it is not representative.
    """
    # read the columns in the BISI configuration file
    bisi_config = read_system_config.read_bisi_config()
    logger.debug(f"BISI = \n {bisi_config.count()}")
    bisi_area_columns = read_system_config.read_yaml_configuration(
        "bisi_area_columns", "global_variables.yaml"
    )

    # calculating BISI is only if the BISI columns are not empty
    if df[bisi_area_columns].isna().all().all():
        return False

    # Prepare the BISI template for output
    output_path = read_system_config.read_yaml_configuration(
        "calculated_bisi", "global_variables.yaml"
    )
    shutil.copyfile(
        read_system_config.read_yaml_configuration(
            "template_bisi", "global_variables.yaml"
        ),
        output_path,
    )
    writer = pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
    )
    wb = load_workbook(output_path)

    # create emtpty return dataframe
    df_density_agg_stdev_nsample = pd.DataFrame()

    # loop over BISI_gebieden to fill BISI table
    for bisi_column in bisi_area_columns:
        bisi_areas = df[bisi_column].dropna().unique().tolist()
        for bisi_area in bisi_areas:
            # filter the data for the BISI area
            df_by_area = df[df[bisi_column].isin([bisi_area])]

            # if year is not specified, use the most recent year for which there is data.
            if year is None:
                year = df_by_area["Monsterjaar"].max()

            # filter the data for the year
            df_by_area_year = df_by_area[df_by_area["Monsterjaar"].isin([year])]
            if df_by_area_year.empty:
                logger.warning(
                    f"Geen data om de BISI index te berekenen voor {bisi_column}:{bisi_area} in {year}."
                )
                break

            # retrieve the BISI-sheet configuration for the selected BISI area, so we know what to look for
            try:
                # try to retrieve the BISI configs
                bisi_sheet_sheetname = bisi_config[
                    bisi_config["BISI_indeling"] == bisi_area
                ]["Sheet_name"].values[0]
                bisi_sheet_area_title = bisi_config[
                    bisi_config["BISI_indeling"] == bisi_area
                ]["BISI_row"].values[0]
                bisi_sheet_sampling_devices = bisi_config[
                    bisi_config["BISI_indeling"] == bisi_area
                ]["Sample_devices"].values[0]

                # transform BISI sampling devices mapping to dictionary
                items = bisi_sheet_sampling_devices.strip("{}").split(", ")
                bisi_sheet_sampling_devices_dict = {}
                for item in items:
                    key, value = item.split(": ")
                    bisi_sheet_sampling_devices_dict[key.strip("'")] = value.strip("'")

                logger.info(
                    f"BISI wordt berekend voor {bisi_column}: {bisi_area} voor het jaar {year}."
                )
            except Exception:
                logger.warning(
                    "Er kan geen BISI configuratie (tabblad naam/titel gebied/bemonsteringsapparaat)"
                    f"worden opgehaald voor {bisi_area} dus er is geen BISI berekend."
                )
                continue

            # write the sample devices in the data according to the BISI sheet (see config)
            df_by_area_year = map_sampling_device_to_bisi(
                df_by_area_year, bisi_sheet_sampling_devices_dict
            )

            # loop over the rows and cells in the BISI sheet to find the right place for the selected bisi_area
            wb.active = wb[bisi_sheet_sheetname]

            found_area = False
            for row in wb.active.iter_rows(max_col=1):
                for cell in row:
                    cell_value = str(cell.value)

                    # find the cell where the bisi_area is equal to the bisi_sheet_area_title in the configs
                    if found_area and (cell_value is None or cell_value == "None"):
                        # if the cell is empty, we are done
                        logger.info(
                            f"BISI tabel is ingevuld voor {bisi_sheet_area_title} voor {year}."
                        )
                        break

                    if cell_value is not None and bisi_sheet_area_title in cell_value:
                        logger.debug(f"cell match = {cell_value}")
                        found_area = True
                        cell_index = cell.row

                        df_bisi_criteria = read_bisi_criteria(
                            output_path, bisi_sheet_sheetname, cell_index
                        )

                        # check the taxa in the BISI sheet
                        check_bisi_taxa(df_bisi_criteria, bisi_area)

                        # map taxa to BISI taxa
                        df_by_area_year = map_taxa_to_bisi(
                            df_by_area_year.copy(), df_bisi_criteria
                        )

                        check_required_area(
                            df_bisi=df_by_area_year, bisi_col=bisi_column
                        )

                        # calculate the average density and standard deviation
                        df_density_agg_stdev = bisi_calculations(
                            df_by_area_year, df_bisi_criteria
                        )

                        # check the number of samples and species in the BISI sheet
                        df_monsters = check_sample_species(
                            df_by_area_year,
                            df_bisi_criteria,
                            bisi_column,
                        )

                        # merge the number of samples with the average density and standard deviation
                        df_density_agg_stdev_nsample = pd.merge(
                            df_density_agg_stdev,
                            df_monsters[["Bemonsteringsapp", "Nmonsters"]],
                            how="left",
                            on=["Bemonsteringsapp"],
                        )

                        # order the rows by the position in the BISI sheet
                        df_density_agg_stdev_nsample.sort_values(
                            by=["Position"], inplace=True
                        )

                        # write the average density, standard deviation an n samples to the bisi sheet
                        df_density_agg_stdev_nsample[
                            ["Nmonsters", "Dichtheid_Aantal", "Stdev"]
                        ].to_excel(
                            writer,
                            sheet_name=bisi_sheet_sheetname,
                            header=None,
                            startcol=12,
                            startrow=cell_index + 1,
                            index=False,
                        )

                        # write the year to the bisi sheet
                        year_df = pd.DataFrame({year})
                        year_df.to_excel(
                            writer,
                            sheet_name=bisi_sheet_sheetname,
                            header=None,
                            startcol=0,
                            startrow=cell_index - 2,
                            index=False,
                        )

                else:  # continue if the inner loop wasn't broken
                    continue
                break
    writer.close()
    msg = "BISI: gereed"
    logger.info(msg)
    print(msg)
    return df_density_agg_stdev_nsample
